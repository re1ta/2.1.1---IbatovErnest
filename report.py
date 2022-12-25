import openpyxl
from openpyxl.styles import Side, Border, Font


class Report:
    def __init__(self, salary_by_year: dict, vacancies_by_year: dict, salary_by_year_for_profession: dict,
                 vacancies_by_year_for_profession: dict, salary_by_city: dict, vacancies_by_city: dict,
                 profession_name: str):
        self.__salary_by_year = salary_by_year
        self.__vacancies_by_year = vacancies_by_year
        self.__salary_by_year_for_profession = salary_by_year_for_profession
        self.__vacancies_by_year_for_profession = vacancies_by_year_for_profession
        self.__salary_by_city = salary_by_city
        self.__vacancies_by_city = vacancies_by_city
        self.__profession_name = profession_name

    def generate_excel(self):
        excel_file = openpyxl.Workbook()
        excel_file.remove(excel_file["Sheet"])
        excel_file.create_sheet("Статистика по годам")
        years = [2007, 2008, 2009, 2010, 2011, 2012, 2013, 2014, 2015, 2016, 2017, 2018, 2019, 2020, 2021, 2022]
        title1 = ["Год", "Средняя зарплата", f"Средняя зарплата - {self.__profession_name}", "Количество вакансий",
                  f"Количество вакансий - {self.__profession_name}"]
        excel_file.worksheets[0].append(title1)
        for i in years:
            if i in self.__salary_by_year.keys():
                excel_file.worksheets[0].append([i, self.__salary_by_year[i], self.__salary_by_year_for_profession[i],
                                                 self.__vacancies_by_year[i],
                                                 self.__vacancies_by_year_for_profession[i]])

        for i in range(len(title1)):  excel_file.worksheets[0].cell(1, i + 1).font = Font(bold=True)

        side = Side(border_style='thin', color="FF000000")
        border = Border(left=side, right=side, top=side, bottom=side)
        for i in range(len(self.__salary_by_year.keys()) + 1):
            for j in range(len(title1)): excel_file.worksheets[0].cell(i + 1, j + 1).border = border

        dims = {}
        for row in excel_file.worksheets[0].rows:
            for cell in row:
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value)) + 2))
        for col, value in dims.items(): excel_file.worksheets[0].column_dimensions[col].width = value

        excel_file.create_sheet("Статистика по городам")
        t2 = ["Город", "Уровень зарплат", "", "Город", "Доля вакансий"]
        excel_file.worksheets[1].append(t2)
        c1 = list(self.__salary_by_city.keys())
        c2 = list(self.__vacancies_by_city.keys())
        for i in range(len(c1)):
            excel_file.worksheets[1].append([c1[i], self.__salary_by_city[c1[i]], "",
                                             c2[i], self.__vacancies_by_city[c2[i]]])
        for i in range(len(t2)): excel_file.worksheets[1].cell(1, i + 1).font = Font(bold=True)

        for i in range(len(c1) + 1):
            for j in range(len(t2)):
                excel_file.worksheets[1].cell(i + 1, j + 1).border = border

        for i in range(2, len(c2) + 2):
            excel_file.worksheets[1].cell(i, 5).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[10]

        dims = {}
        for i in excel_file.worksheets[1].rows:
            for j in i:
                if j.value: dims[j.column_letter] = max((dims.get(j.column_letter, 0), len(str(j.value)) + 2))
        for i, j in dims.items():
            excel_file.worksheets[1].column_dimensions[i].width = j

        excel_file.save("report.xlsx")
